import json
import os
import re
import uuid
from io import BytesIO

from flask import Flask, render_template, request, send_file, session, after_this_request

from test_generator import generate_test_cases, COLUMNS

app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key-change-in-production")


def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/", methods=["GET"])
def index():
    @after_this_request
    def apply_no_cache(response):
        return no_cache(response)
    return render_template("index.html", prefill=[])


@app.route("/generate", methods=["POST"])
def generate():
    @after_this_request
    def apply_no_cache(response):
        return no_cache(response)

    req_names   = request.form.getlist("req_name[]")
    req_details = request.form.getlist("requirements[]")

    pairs = []
    for name, detail in zip(req_names, req_details):
        name   = name.strip()
        detail = detail.strip()
        if name and detail:
            pairs.append({"req_name": name, "requirements": detail})

    if not pairs:
        return render_template(
            "index.html",
            error="Please provide at least one Requirement Name and Requirement Details.",
            prefill=[{"req_name": n.strip(), "requirements": d.strip()}
                     for n, d in zip(req_names, req_details)],
        )

    all_requirements = []
    used_ai = False
    fallback_reason = None
    for pair in pairs:
        try:
            cases = generate_test_cases(pair["req_name"], pair["requirements"])
        except Exception as exc:
            return render_template(
                "index.html",
                error=f"Could not generate test cases for '{pair['req_name']}': {exc}",
                prefill=pairs,
            )

        if cases and cases[0].get("_generation_source") == "ai":
            used_ai = True
        elif cases and cases[0].get("_fallback_reason") and not fallback_reason:
            fallback_reason = cases[0]["_fallback_reason"]

        # Strip internal bookkeeping keys before they reach the template/Excel payload
        for c in cases:
            c.pop("_generation_source", None)
            c.pop("_fallback_reason", None)

        all_requirements.append({
            "req_name":     pair["req_name"],
            "requirements": pair["requirements"],
            "test_cases":   cases,
        })

    total_cases = sum(len(r["test_cases"]) for r in all_requirements)

    # Embed data in the page itself — no session dependency for download
    payload_json = json.dumps(all_requirements)

    generation_notice = None
    if not used_ai:
        generation_notice = (
            "Generated using the built-in rule engine"
            + (f" ({fallback_reason})." if fallback_reason else ".")
        )

    return render_template(
        "results.html",
        all_requirements=all_requirements,
        total_cases=total_cases,
        columns=COLUMNS,
        payload_json=payload_json,
        generation_notice=generation_notice,
    )


@app.route("/download", methods=["POST"])
def download():
    """
    Receives the full test-case payload as JSON in the POST body,
    builds an Excel workbook, and returns it as a download.
    No session or server-side cache required.
    """
    try:
        payload_str = request.form.get("payload", "")
        if not payload_str:
            return "No data received. Please generate test cases first.", 400

        payload_str = payload_str.strip()
        all_requirements = json.loads(payload_str)

        if not isinstance(all_requirements, list) or not all_requirements:
            return "Empty or invalid payload.", 400

        # Validate each requirement has test_cases
        for req in all_requirements:
            if "test_cases" not in req:
                req["test_cases"] = []

    except json.JSONDecodeError as exc:
        return f"JSON parse error: {exc}. Payload length: {len(payload_str)}", 400
    except Exception as exc:
        return f"Payload error: {exc}", 400

    from openpyxl import Workbook
    from openpyxl.styles import Font as XLFont, Alignment as XLAlign, PatternFill as XLFill
    from openpyxl.utils import get_column_letter as gcl

    wb = Workbook()
    used_sheet_names: set = set()

    # ── Unique sheet name helper ───────────────────────────────────────
    def unique_sheet_name(raw: str) -> str:
        clean = re.sub(r"[\\/*?\[\]:]", "", raw)[:28].strip() or "Sheet"
        name, counter = clean, 2
        while name in used_sheet_names:
            suffix = f" {counter}"
            name = clean[: 31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(name)
        return name

    # ── Cover sheet ───────────────────────────────────────────────────
    cover_ws = wb.active
    cover_ws.title = "Cover Page"
    cover_ws.sheet_view.showGridLines = False
    used_sheet_names.add("Cover Page")

    for col in range(1, 6):
        cover_ws.column_dimensions[gcl(col)].width = 28

    cover_ws.row_dimensions[2].height = 50
    t = cover_ws.cell(row=2, column=1, value="TEST CASE DOCUMENT")
    t.font      = XLFont(name="Calibri", bold=True, size=22, color="1F2A44")
    t.alignment = XLAlign(horizontal="left", vertical="center")
    cover_ws.merge_cells("A2:E2")

    cover_ws.row_dimensions[3].height = 6
    for col in range(1, 6):
        c = cover_ws.cell(row=3, column=col)
        c.fill = XLFill(start_color="4DD0C4", end_color="4DD0C4", fill_type="solid")

    def make_label(ws, row, text):
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=1, value=text)
        c.font      = XLFont(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = XLFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
        c.alignment = XLAlign(horizontal="left", vertical="center", indent=1)

    def make_value(ws, row, text):
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=2, value=str(text))
        c.font      = XLFont(name="Calibri", size=11, color="1F2A44")
        c.alignment = XLAlign(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.merge_cells(f"B{row}:E{row}")

    total_cases = sum(len(r["test_cases"]) for r in all_requirements)
    make_label(cover_ws, 5, "Requirements");     make_value(cover_ws, 5, ", ".join(r["req_name"] for r in all_requirements))
    make_label(cover_ws, 6, "Total Test Cases"); make_value(cover_ws, 6, total_cases)
    make_label(cover_ws, 7, "Document Status");  make_value(cover_ws, 7, "Draft")

    # ── Per-requirement test case breakdown ───────────────────────────
    # Blank spacer row
    cover_ws.row_dimensions[8].height = 12

    # Section heading
    cover_ws.row_dimensions[9].height = 22
    sec = cover_ws.cell(row=9, column=1, value="Test Cases per Requirement")
    sec.font      = XLFont(name="Calibri", bold=True, size=12, color="1F2A44")
    sec.alignment = XLAlign(horizontal="left", vertical="center")
    cover_ws.merge_cells("A9:E9")

    # Teal separator
    cover_ws.row_dimensions[10].height = 4
    for ci in range(1, 6):
        cover_ws.cell(row=10, column=ci).fill = XLFill(start_color="4DD0C4", end_color="4DD0C4", fill_type="solid")

    # Column headers for the breakdown table
    cover_ws.row_dimensions[11].height = 22
    for col_idx, heading in enumerate(["Requirement Name", "Test Cases Generated"], start=1):
        ch = cover_ws.cell(row=11, column=col_idx, value=heading)
        ch.font      = XLFont(name="Calibri", bold=True, size=10, color="FFFFFF")
        ch.fill      = XLFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
        ch.alignment = XLAlign(horizontal="left" if col_idx == 1 else "center", vertical="center", indent=1)

    # One row per requirement
    ALT_FILL = XLFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
    for i, req in enumerate(all_requirements):
        row_no = 12 + i
        cover_ws.row_dimensions[row_no].height = 22

        fill = ALT_FILL if i % 2 == 0 else None

        name_cell = cover_ws.cell(row=row_no, column=1, value=req["req_name"])
        name_cell.font      = XLFont(name="Calibri", size=10, color="1F2A44")
        name_cell.alignment = XLAlign(horizontal="left", vertical="center", indent=1)
        if fill:
            name_cell.fill = fill

        count_cell = cover_ws.cell(row=row_no, column=2, value=len(req["test_cases"]))
        count_cell.font      = XLFont(name="Calibri", bold=True, size=10, color="1F2A44")
        count_cell.alignment = XLAlign(horizontal="center", vertical="center")
        if fill:
            count_cell.fill = fill

    # Total row
    total_row = 12 + len(all_requirements)
    cover_ws.row_dimensions[total_row].height = 24
    tl = cover_ws.cell(row=total_row, column=1, value="TOTAL")
    tl.font      = XLFont(name="Calibri", bold=True, size=10, color="FFFFFF")
    tl.fill      = XLFill(start_color="4DD0C4", end_color="4DD0C4", fill_type="solid")
    tl.alignment = XLAlign(horizontal="left", vertical="center", indent=1)

    tv = cover_ws.cell(row=total_row, column=2, value=total_cases)
    tv.font      = XLFont(name="Calibri", bold=True, size=10, color="FFFFFF")
    tv.fill      = XLFill(start_color="4DD0C4", end_color="4DD0C4", fill_type="solid")
    tv.alignment = XLAlign(horizontal="center", vertical="center")

    # ── One sheet per requirement ──────────────────────────────────────
    COL_WIDTHS = {
        "Sl. No.": 8, "TC ID": 12, "TC Objective": 32,
        "TC Description": 42, "TC Steps": 46, "Test Data": 26,
        "Expected Output": 36, "Requirement ID": 15,
        "Priority": 11, "Status": 14, "Pass/Fail": 12,
    }

    HDR_FILL  = XLFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    HDR_FONT  = XLFont(name="Calibri", bold=True, color="FFFFFF", size=10)
    HDR_ALIGN = XLAlign(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGN = XLAlign(wrap_text=True, vertical="top")

    for req in all_requirements:
        sheet_name = unique_sheet_name(req["req_name"])
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = True

        # ── Requirement info rows at the top ──────────────────────────
        # Row 1: Requirement Name
        rn_label = ws.cell(row=1, column=1, value="Requirement Name")
        rn_label.font      = XLFont(name="Calibri", bold=True, size=10, color="FFFFFF")
        rn_label.fill      = XLFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
        rn_label.alignment = XLAlign(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 22

        rn_val = ws.cell(row=1, column=2, value=req["req_name"])
        rn_val.font      = XLFont(name="Calibri", bold=True, size=11, color="1F2A44")
        rn_val.alignment = XLAlign(horizontal="left", vertical="center", indent=1)
        ws.merge_cells("B1:K1")

        # Row 2: Requirement Details
        rd_label = ws.cell(row=2, column=1, value="Requirement Details")
        rd_label.font      = XLFont(name="Calibri", bold=True, size=10, color="FFFFFF")
        rd_label.fill      = XLFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
        rd_label.alignment = XLAlign(horizontal="left", vertical="top", indent=1)
        ws.row_dimensions[2].height = 40

        rd_val = ws.cell(row=2, column=2, value=req["requirements"])
        rd_val.font      = XLFont(name="Calibri", size=10, color="333333")
        rd_val.alignment = XLAlign(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws.merge_cells("B2:K2")

        # Row 3: blank spacer
        ws.row_dimensions[3].height = 6
        for ci in range(1, 12):
            ws.cell(row=3, column=ci).fill = XLFill(start_color="4DD0C4", end_color="4DD0C4", fill_type="solid")

        # Row 4: column headers
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = HDR_ALIGN
            ws.column_dimensions[gcl(col_idx)].width = COL_WIDTHS.get(col_name, 20)
        ws.row_dimensions[4].height = 28

        # Freeze below the header row
        ws.freeze_panes = "A5"

        # ── Data rows ─────────────────────────────────────────────────
        for row_idx, tc in enumerate(req["test_cases"], start=5):
            sl_no = row_idx - 4   # resets to 1 for every requirement sheet
            row_vals = [
                tc.get("Sl. No.",         sl_no),
                tc.get("TC ID",           f"TC-{sl_no:03d}"),
                tc.get("TC Objective",    ""),
                tc.get("TC Description",  ""),
                tc.get("TC Steps",        ""),
                tc.get("Test Data",       ""),
                tc.get("Expected Output", ""),
                tc.get("Requirement ID",  ""),
                tc.get("Priority",        ""),
                tc.get("Status",          "Not Executed"),
                tc.get("Pass/Fail",       "N/A"),
            ]
            for col_idx, cell_val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                cell.alignment = CELL_ALIGN

            steps_text = str(tc.get("TC Steps", ""))
            lines = max(steps_text.count("\n") + 1, 1)
            ws.row_dimensions[row_idx].height = max(20, min(lines * 15, 120))

    # ── Save & return ──────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", all_requirements[0]["req_name"]).strip("_") or "requirements"
    filename = f"{safe_name}_TestCases.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=5000)
